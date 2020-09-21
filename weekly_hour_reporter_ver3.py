import pandas as pd
import matplotlib.pyplot as plt
import openpyxl
import dropbox
import os
sheet_data_frame = [0 for i in range(100)]
plt.close('all')
pd.set_option('display.expand_frame_repr', False)
from matplotlib import style
plt.style.use('seaborn')


def main_menu():
    print('Lista uzytkownikow:')
    global user_list
    user_list = ['1. user1','2. user2','3. user3','4. user4',
                 '5. user5','6. user6','7. user7']
    initial = ['LD', 'TS', 'PM', 'MM', 'PZ', 'PS', 'DM']
    for user_name in user_list:
        print(user_name)
    user_id = input('Prosze wybrac numer z listy: \n')
    return user_id, initial[int(user_id)-1]

def dropbox_connect(arg1):
    user_id = arg1[0]
    initial = arg1[1]
    indeks = int(user_id) - 1
    global path_folder
    path_folder = ['/user1 (1)/','/user2 (1)/','/user3 (1)/','/user4 (1)/',
                    '/user5 (1)/','/user6 (1)/','/user7 (1)/']
    #global sheet_name
    sheet_date = ['012019', '022019', '032019', '042019', '052019', '062019', '072019', '082019', '092019',
                  '102019', '112019', '122019']
    imiona = ['_LUKASZ_', '_TOMASZ_', '_PAWEL_', '_MACIEJ_', '_PRZEMYSLAW_', '_PRZEMYSLAW_', '_DANIEL_']
    nazwiska = ['user1', 'user2', 'user3', 'user4', 'user5', 'user6', 'user7']
    #sheet_name1 = str(initial) + '012019' + imiona[indeks] + nazwiska[indeks] + '.xlsm'
    sheet_name = []
    for date in sheet_date:
        sheet_name.append(str(initial) + date + imiona[indeks] + nazwiska[indeks] + '.xlsm')

    #TS012019_TOMASZ_SOLAREWICZ.xlsm
    global app_key
    app_key = '#app-key#'
    global app_secret
    app_secret = '#app-secret#'
    global token
    token = ['1','user2_token', '3', '4', '5', '6', '7']
    dbx = dropbox.Dropbox(token[indeks]) #TODO zmienic tutaj po uzupelnieniu listy tokenow
    path_for_pandas = []
    #print(dbx.users_get_current_account())
    print('Connecting with dropbox...')
    for entry in dbx.files_list_folder(path_folder[indeks]).entries:
        for s_n in sheet_name:
            if entry.name == s_n:  # sprawdza czy na dropie jest poprawnie nazwany plik
                full_path_to_dropbox = path_folder[indeks] + s_n
                full_path_to_local = 'C:\Timesheet_old\\' + initial + '\\' + s_n
                print('Downloading file from dropbox...')
                print(full_path_to_local)
                if os.path.exists('C:\Timesheet_old\\' + initial + '\\') != True:
                    os.mkdir('C:\Timesheet_old\\' + initial + '\\')
                dbx.files_download_to_file(full_path_to_local, full_path_to_dropbox)
                path_for_pandas.append(full_path_to_local)
    #print('Tworzenie kopii zapasowej dla tygodnia {}...'.format(which_week))
    return path_for_pandas

def define_start_stop(arg1):
    path_for_pandas = arg1
    """
    This function will look in what excel file user's input week is found.
    """
    global start, stop
    start = input("Please input on which week you want to start:\n")
    stop = input("Please input on which week do you want to stop:\n")
    start_int = start[-2:]
    stop_int = stop[-2:]
    dict_of_excels = {}
    returned_dict = {}
    temp_dict = {}
    try:
        wb01 = openpyxl.load_workbook(path_for_pandas[0], read_only=True)
        dict_of_excels['wb01'] = wb01
        wb02 = openpyxl.load_workbook(path_for_pandas[1], read_only=True)
        dict_of_excels['wb02'] = wb02
        wb03 = openpyxl.load_workbook(path_for_pandas[2], read_only=True)
        dict_of_excels['wb03'] = wb03
        wb04 = openpyxl.load_workbook(path_for_pandas[3], read_only=True)
        dict_of_excels['wb04'] = wb04
        wb05 = openpyxl.load_workbook(path_for_pandas[4], read_only=True)
        dict_of_excels['wb05'] = wb05
        wb06 = openpyxl.load_workbook(path_for_pandas[5], read_only=True)
        dict_of_excels['wb06'] = wb06
        wb07 = openpyxl.load_workbook(path_for_pandas[6], read_only=True)
        dict_of_excels['wb07'] = wb07
        wb08 = openpyxl.load_workbook(path_for_pandas[7], read_only=True)
        dict_of_excels['wb08'] = wb08
        wb09 = openpyxl.load_workbook(path_for_pandas[8], read_only=True)
        dict_of_excels['wb09'] = wb09
        wb10 = openpyxl.load_workbook(path_for_pandas[9], read_only=True)
        dict_of_excels['wb10'] = wb10
        wb11 = openpyxl.load_workbook(path_for_pandas[10], read_only=True)
        dict_of_excels['wb11'] = wb11
        wb12 = openpyxl.load_workbook(path_for_pandas[11], read_only=True)
        dict_of_excels['wb12'] = wb12
    except:
        pass

    for sheet in dict_of_excels:
        temp_dict[sheet] = dict_of_excels[sheet].sheetnames
        for i in range(int(start_int), (int(stop_int) + 1)):  # we are looping through weeks
            if i < 10:
                week_for_sheet = 'W-0' + str(i)  # this brings index value of sheet
                if week_for_sheet in temp_dict[sheet]:
                    sheet_index = dict_of_excels[sheet].worksheets.index(dict_of_excels[sheet][week_for_sheet])
                    try:
                        if dict_of_excels[sheet] == dict_of_excels['wb01']:
                            which_sheet = 'wb01'
                        if dict_of_excels[sheet] == dict_of_excels['wb02']:
                            which_sheet = 'wb02'
                        if dict_of_excels[sheet] == dict_of_excels['wb03']:
                            which_sheet = 'wb03'
                        if dict_of_excels[sheet] == dict_of_excels['wb04']:
                            which_sheet = 'wb04'
                        if dict_of_excels[sheet] == dict_of_excels['wb05']:
                            which_sheet = 'wb05'
                        if dict_of_excels[sheet] == dict_of_excels['wb06']:
                            which_sheet = 'wb06'
                        if dict_of_excels[sheet] == dict_of_excels['wb07']:
                            which_sheet = 'wb07'
                        if dict_of_excels[sheet] == dict_of_excels['wb08']:
                            which_sheet = 'wb08'
                        if dict_of_excels[sheet] == dict_of_excels['wb09']:
                            which_sheet = 'wb09'
                        if dict_of_excels[sheet] == dict_of_excels['wb10']:
                            which_sheet = 'wb10'
                        if dict_of_excels[sheet] == dict_of_excels['wb11']:
                            which_sheet = 'wb11'
                        if dict_of_excels[sheet] == dict_of_excels['wb12']:
                            which_sheet = 'wb12'
                    except:
                        pass
                    if week_for_sheet not in returned_dict.keys():
                        returned_dict[week_for_sheet] = sheet_index, which_sheet
                    else:
                        returned_dict[week_for_sheet + '_1'] = sheet_index, which_sheet
            else:
                week_for_sheet = 'W-' + str(i)  # this brings index value of sheet
                if week_for_sheet in temp_dict[sheet]:
                    sheet_index = dict_of_excels[sheet].worksheets.index(dict_of_excels[sheet][week_for_sheet])
                    try:
                        if dict_of_excels[sheet] == dict_of_excels['wb01']:
                            which_sheet = 'wb01'
                        if dict_of_excels[sheet] == dict_of_excels['wb02']:
                            which_sheet = 'wb02'
                        if dict_of_excels[sheet] == dict_of_excels['wb03']:
                            which_sheet = 'wb03'
                        if dict_of_excels[sheet] == dict_of_excels['wb04']:
                            which_sheet = 'wb04'
                        if dict_of_excels[sheet] == dict_of_excels['wb05']:
                            which_sheet = 'wb05'
                        if dict_of_excels[sheet] == dict_of_excels['wb06']:
                            which_sheet = 'wb06'
                        if dict_of_excels[sheet] == dict_of_excels['wb07']:
                            which_sheet = 'wb07'
                        if dict_of_excels[sheet] == dict_of_excels['wb08']:
                            which_sheet = 'wb08'
                        if dict_of_excels[sheet] == dict_of_excels['wb09']:
                            which_sheet = 'wb09'
                        if dict_of_excels[sheet] == dict_of_excels['wb10']:
                            which_sheet = 'wb10'
                        if dict_of_excels[sheet] == dict_of_excels['wb11']:
                            which_sheet = 'wb11'
                        if dict_of_excels[sheet] == dict_of_excels['wb12']:
                            which_sheet = 'wb12'
                    except:
                        pass
                    if week_for_sheet not in returned_dict.keys():
                        returned_dict[week_for_sheet] = sheet_index, which_sheet
                    else:
                        returned_dict[week_for_sheet + '_1'] = sheet_index, which_sheet
    print(returned_dict)
    return returned_dict



def show_values(axis):
    axis.set_xlabel("Client")
    axis.set_ylabel("Amount of hours")
    for p in axis.patches:
        width, height = p.get_width(), p.get_height()
        x, y = p.get_xy()
        axis.annotate((height), (p.get_x() + 0.375 * width, p.get_y() + 0.5 * height))


def fill(Df):
    Df.fillna(value={"REGULAR WORKING TIME": 0, "ORDERED OVERTIME": 0, "TOTAL PROJECT HOURS": 0, "TASK": "",
                  "CUSTOMER": "", "PROJECT NUMBER/ORDER NUMBER": ""}, inplace=True)
    Df.ffill(0, inplace=True)
    Df["DATETIME"] = Df["DATETIME"].apply(lambda x: str(x.isocalendar()[0]) + '-' +
                                                      str(x.isocalendar()[1]).zfill(2))

def create_dataframe(arg1):
    path_for_pandas = arg1
    i=0
    try:
        wb01 = path_for_pandas[0]
        wb02 = path_for_pandas[1]
        wb03 = path_for_pandas[2]
        wb04 = path_for_pandas[3]
        wb05 = path_for_pandas[4]
        wb06 = path_for_pandas[5]
        wb07 = path_for_pandas[6]
        wb08 = path_for_pandas[7]
        wb09 = path_for_pandas[8]
        wb10 = path_for_pandas[9]
        wb11 = path_for_pandas[10]
        wb12 = path_for_pandas[11]
    except:
        pass
    Df1 = pd.DataFrame
    Df2 = pd.DataFrame
    Df3 = pd.DataFrame
    Df4 = pd.DataFrame
    Df5 = pd.DataFrame
    Df6 = pd.DataFrame
    Df7 = pd.DataFrame
    Df8 = pd.DataFrame
    Df9 = pd.DataFrame
    Df10 = pd.DataFrame
    Df11 = pd.DataFrame
    Df12 = pd.DataFrame
    Df13 = pd.DataFrame
    Df14 = pd.DataFrame
    Df15 = pd.DataFrame
    Df16 = pd.DataFrame
    Df17 = pd.DataFrame
    Df18 = pd.DataFrame
    Df19 = pd.DataFrame
    Df20 = pd.DataFrame
    Df21 = pd.DataFrame
    Df22 = pd.DataFrame
    Df23 = pd.DataFrame
    Df24 = pd.DataFrame
    Df25 = pd.DataFrame
    Df26 = pd.DataFrame
    Df27 = pd.DataFrame
    Df28 = pd.DataFrame
    Df29 = pd.DataFrame
    Df30 = pd.DataFrame
    Df31 = pd.DataFrame
    Df32 = pd.DataFrame
    Df33 = pd.DataFrame
    Df34 = pd.DataFrame
    Df35 = pd.DataFrame
    Df36 = pd.DataFrame
    Df37 = pd.DataFrame
    Df38 = pd.DataFrame
    Df39 = pd.DataFrame
    Df40 = pd.DataFrame
    Df41 = pd.DataFrame
    Df42 = pd.DataFrame
    Df43 = pd.DataFrame
    Df44 = pd.DataFrame
    Df45 = pd.DataFrame
    Df46 = pd.DataFrame
    Df47 = pd.DataFrame
    Df48 = pd.DataFrame
    Df49 = pd.DataFrame
    Df50 = pd.DataFrame
    Df51 = pd.DataFrame
    Df52 = pd.DataFrame
    Df53 = pd.DataFrame
    Df54 = pd.DataFrame
    returned_dictionary = define_start_stop(path_for_pandas)
    klucze = returned_dictionary.keys()
    for klucz in klucze:
        if returned_dictionary[klucz][1] == 'wb01':
            excel_file = wb01
        elif returned_dictionary[klucz][1] == 'wb02':
            excel_file = wb02
        elif returned_dictionary[klucz][1] == 'wb03':
            excel_file = wb03
        elif returned_dictionary[klucz][1] == 'wb04':
            excel_file = wb04
        elif returned_dictionary[klucz][1] == 'wb05':
            excel_file = wb05
        elif returned_dictionary[klucz][1] == 'wb06':
            excel_file = wb06
        elif returned_dictionary[klucz][1] == 'wb07':
            excel_file = wb07
        elif returned_dictionary[klucz][1] == 'wb08':
            excel_file = wb08
        elif returned_dictionary[klucz][1] == 'wb09':
            excel_file = wb09
        elif returned_dictionary[klucz][1] == 'wb10':
            excel_file = wb10
        elif returned_dictionary[klucz][1] == 'wb11':
            excel_file = wb11
        elif returned_dictionary[klucz][1] == 'wb12':
            excel_file = wb12
        sheet_data_frame[i] = pd.read_excel(excel_file, sheet_name=returned_dictionary[klucz][0], header=0, skipfooter=2, usecols=range(0, 8), names=["DATETIME", "DAY", "PROJECT NUMBER/ORDER NUMBER","CUSTOMER", "TASK", "REGULAR WORKING TIME", "ORDERED OVERTIME", "TOTAL PROJECT HOURS"])
        i += 1
    dataframes = []

    try:
        Df1 = sheet_data_frame[0]
        if not Df1.empty:
            dataframes.append(Df1)
        Df2 = sheet_data_frame[1]
        if not Df2.empty:
            dataframes.append(Df2)
        Df3 = sheet_data_frame[2]
        if not Df3.empty:
            dataframes.append(Df3)
        Df4 = sheet_data_frame[3]
        if not Df4.empty:
            dataframes.append(Df4)
        Df5 = sheet_data_frame[4]
        if not Df5.empty:
            dataframes.append(Df5)
        Df6 = sheet_data_frame[5]
        if not Df6.empty:
            dataframes.append(Df6)
        Df7 = sheet_data_frame[6]
        if not Df7.empty:
            dataframes.append(Df7)
        Df8 = sheet_data_frame[7]
        if not Df8.empty:
            dataframes.append(Df8)
        Df9 = sheet_data_frame[8]
        if not Df9.empty:
            dataframes.append(Df9)
        Df10 = sheet_data_frame[9]
        if not Df10.empty:
            dataframes.append(Df10)
        Df11 = sheet_data_frame[10]
        if not Df11.empty:
            dataframes.append(Df11)
        Df12 = sheet_data_frame[11]
        if not Df12.empty:
            dataframes.append(Df12)
        Df13 = sheet_data_frame[12]
        if not Df13.empty:
            dataframes.append(Df13)
        Df14 = sheet_data_frame[13]
        if not Df14.empty:
            dataframes.append(Df14)
        Df15 = sheet_data_frame[14]
        if not Df15.empty:
            dataframes.append(Df15)
        Df16 = sheet_data_frame[15]
        if not Df16.empty:
            dataframes.append(Df16)
        Df17 = sheet_data_frame[16]
        if not Df17.empty:
            dataframes.append(Df17)
        Df18 = sheet_data_frame[17]
        if not Df18.empty:
            dataframes.append(Df18)
        Df19 = sheet_data_frame[18]
        if not Df19.empty:
            dataframes.append(Df19)
        Df20 = sheet_data_frame[19]
        if not Df20.empty:
            dataframes.append(Df20)
        Df21 = sheet_data_frame[20]
        if not Df21.empty:
            dataframes.append(Df21)
        Df22 = sheet_data_frame[21]
        if not Df22.empty:
            dataframes.append(Df22)
        Df23 = sheet_data_frame[22]
        if not Df23.empty:
            dataframes.append(Df23)
        Df24 = sheet_data_frame[23]
        if not Df24.empty:
            dataframes.append(Df24)
        Df25 = sheet_data_frame[24]
        if not Df25.empty:
            dataframes.append(Df25)
        Df26 = sheet_data_frame[25]
        if not Df26.empty:
            dataframes.append(Df26)
        Df27 = sheet_data_frame[26]
        if not Df27.empty:
            dataframes.append(Df27)
        Df28 = sheet_data_frame[27]
        if not Df28.empty:
            dataframes.append(Df28)
        Df29 = sheet_data_frame[28]
        if not Df29.empty:
            dataframes.append(Df29)
        Df30 = sheet_data_frame[29]
        if not Df30.empty:
            dataframes.append(Df30)
        Df31 = sheet_data_frame[30]
        if not Df31.empty:
            dataframes.append(Df31)
        Df32 = sheet_data_frame[31]
        if not Df32.empty:
            dataframes.append(Df32)
        Df33 = sheet_data_frame[32]
        if not Df33.empty:
            dataframes.append(Df33)
        Df34 = sheet_data_frame[33]
        if not Df34.empty:
            dataframes.append(Df34)
        Df35 = sheet_data_frame[34]
        if not Df35.empty:
            dataframes.append(Df35)
        Df36 = sheet_data_frame[35]
        if not Df36.empty:
            dataframes.append(Df36)
        Df37 = sheet_data_frame[36]
        if not Df37.empty:
            dataframes.append(Df37)
        Df38 = sheet_data_frame[37]
        if not Df38.empty:
            dataframes.append(Df38)
        Df39 = sheet_data_frame[38]
        if not Df39.empty:
            dataframes.append(Df39)
        Df40 = sheet_data_frame[39]
        if not Df40.empty:
            dataframes.append(Df40)
        Df41 = sheet_data_frame[40]
        if not Df41.empty:
            dataframes.append(Df41)
        Df42 = sheet_data_frame[41]
        if not Df42.empty:
            dataframes.append(Df42)
        Df43 = sheet_data_frame[42]
        if not Df43.empty:
            dataframes.append(Df43)
        Df44 = sheet_data_frame[43]
        if not Df44.empty:
            dataframes.append(Df44)
        Df45 = sheet_data_frame[44]
        if not Df45.empty:
            dataframes.append(Df45)
        Df46 = sheet_data_frame[45]
        if not Df46.empty:
            dataframes.append(Df46)
        Df47 = sheet_data_frame[46]
        if not Df47.empty:
            dataframes.append(Df47)
        Df48 = sheet_data_frame[47]
        if not Df48.empty:
            dataframes.append(Df48)
        Df49 = sheet_data_frame[48]
        if not Df49.empty:
            dataframes.append(Df49)
        Df50 = sheet_data_frame[49]
        if not Df50.empty:
            dataframes.append(Df50)
        Df51 = sheet_data_frame[50]
        if not Df51.empty:
            dataframes.append(Df51)
        Df52 = sheet_data_frame[51]
        if not Df52.empty:
            dataframes.append(Df52)
        Df53 = sheet_data_frame[52]
        if not Df53.empty:
            dataframes.append(Df53)
        Df54 = sheet_data_frame[53]
        if not Df54.empty:
            dataframes.append(Df54)
    except:
        pass


    Df = pd.DataFrame
    try:
        fill(Df1)
        Df1 = Df1[Df1["TOTAL PROJECT HOURS"] != 0]
        title01 = 'W' + Df1['DATETIME'].iloc[0][-2:]
        ax1 = Df1.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title01)
        ax1.grid(color='#988e8c')
        show_values(ax1)

        fill(Df2)
        Df2 = Df2[Df2["TOTAL PROJECT HOURS"] != 0]
        title02 = 'W' + Df2['DATETIME'].iloc[0][-2:]
        ax2 = Df2.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title02)
        ax2.grid(color='#988e8c')
        show_values(ax2)
        fill(Df3)
        Df3 = Df3[Df3["TOTAL PROJECT HOURS"] != 0]
        title03 = 'W' + Df3['DATETIME'].iloc[0][-2:]
        ax3 = Df3.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title03)
        ax3.grid(color='#988e8c')
        show_values(ax3)
        fill(Df4)
        Df4 = Df4[Df4["TOTAL PROJECT HOURS"] != 0]
        title04 = 'W' + Df4['DATETIME'].iloc[0][-2:]
        ax4 = Df4.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title04)
        ax4.grid(color='#988e8c')
        show_values(ax4)
        fill(Df5)
        Df5 = Df5[Df5["TOTAL PROJECT HOURS"] != 0]
        title05 = 'W' + Df5['DATETIME'].iloc[0][-2:]
        ax5 = Df5.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title05)
        ax5.grid(color='#988e8c')
        show_values(ax5)
        fill(Df6)
        Df6 = Df6[Df6["TOTAL PROJECT HOURS"] != 0]
        title06 = 'W' + Df6['DATETIME'].iloc[0][-2:]
        ax6 = Df6.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title06)
        ax6.grid(color='#988e8c')
        show_values(ax6)
        fill(Df7)
        Df7 = Df7[Df7["TOTAL PROJECT HOURS"] != 0]
        title07 = 'W' + Df7['DATETIME'].iloc[0][-2:]
        ax7 = Df7.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title07)
        ax7.grid(color='#988e8c')
        show_values(ax7)

        fill(Df8)
        Df8 = Df8[Df8["TOTAL PROJECT HOURS"] != 0]
        title08 = 'W' + Df8['DATETIME'].iloc[0][-2:]
        ax8 = Df8.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title08)
        ax8.grid(color='#988e8c')
        show_values(ax8)

        fill(Df9)
        Df9 = Df9[Df9["TOTAL PROJECT HOURS"] != 0]
        title09= 'W' + Df9['DATETIME'].iloc[0][-2:]
        ax9 = Df9.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title09)
        ax9.grid(color='#988e8c')
        show_values(ax9)

        fill(Df10)
        Df10 = Df10[Df10["TOTAL PROJECT HOURS"] != 0]
        title10 = 'W' + Df10['DATETIME'].iloc[0][-2:]
        ax10 = Df10.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title10)
        ax10.grid(color='#988e8c')
        show_values(ax10)

        fill(Df11)
        Df11 = Df11[Df11["TOTAL PROJECT HOURS"] != 0]
        title11 = 'W' + Df11['DATETIME'].iloc[0][-2:]
        ax11 = Df11.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title11)
        ax11.grid(color='#988e8c')
        show_values(ax11)

        fill(Df12)
        Df12 = Df12[Df12["TOTAL PROJECT HOURS"] != 0]
        title12 = 'W' + Df12['DATETIME'].iloc[0][-2:]
        ax12 = Df12.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title12)
        ax12.grid(color='#988e8c')
        show_values(ax12)

        fill(Df13)
        Df13 = Df13[Df13["TOTAL PROJECT HOURS"] != 0]
        title13 = 'W' + Df13['DATETIME'].iloc[0][-2:]
        ax13 = Df13.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title13)
        ax13.grid(color='#988e8c')
        show_values(ax13)

        fill(Df14)
        Df14 = Df14[Df14["TOTAL PROJECT HOURS"] != 0]
        title14 = 'W' + Df14['DATETIME'].iloc[0][-2:]
        ax14 = Df14.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title14)
        ax14.grid(color='#988e8c')
        show_values(ax14)

        fill(Df15)
        Df15 = Df15[Df15["TOTAL PROJECT HOURS"] != 0]
        title15 = 'W' + Df15['DATETIME'].iloc[0][-2:]
        ax15 = Df15.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title15)
        ax15.grid(color='#988e8c')
        show_values(ax15)

        fill(Df16)
        Df16 = Df16[Df16["TOTAL PROJECT HOURS"] != 0]
        title16 = 'W' + Df16['DATETIME'].iloc[0][-2:]
        ax16= Df16.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title16)
        ax16.grid(color='#988e8c')
        show_values(ax16)

        fill(Df17)
        Df17 = Df17[Df17["TOTAL PROJECT HOURS"] != 0]
        title17 = 'W' + Df17['DATETIME'].iloc[0][-2:]
        ax17 = Df17.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title17)
        ax17.grid(color='#988e8c')
        show_values(ax17)

        fill(Df18)
        Df18 = Df18[Df18["TOTAL PROJECT HOURS"] != 0]
        title18 = 'W' + Df18['DATETIME'].iloc[0][-2:]
        ax18 = Df18.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title18)
        ax18.grid(color='#988e8c')
        show_values(ax18)

        fill(Df19)
        Df19 = Df19[Df19["TOTAL PROJECT HOURS"] != 0]
        title19 = 'W' + Df19['DATETIME'].iloc[0][-2:]
        ax19 = Df19.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title19)
        ax19.grid(color='#988e8c')
        show_values(ax19)

        fill(Df20)
        Df20 = Df20[Df20["TOTAL PROJECT HOURS"] != 0]
        title20 = 'W' + Df20['DATETIME'].iloc[0][-2:]
        ax20 = Df20.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title20)
        ax20.grid(color='#988e8c')
        show_values(ax20)

        fill(Df21)
        Df21 = Df21[Df21["TOTAL PROJECT HOURS"] != 0]
        title21 = 'W' + Df21['DATETIME'].iloc[0][-2:]
        ax21 = Df21.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title21)
        ax21.grid(color='#988e8c')
        show_values(ax21)

        fill(Df22)
        Df22 = Df22[Df22["TOTAL PROJECT HOURS"] != 0]
        title22 = 'W' + Df22['DATETIME'].iloc[0][-2:]
        ax22 = Df22.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title22)
        ax22.grid(color='#988e8c')
        show_values(ax22)

        fill(Df23)
        Df23 = Df23[Df23["TOTAL PROJECT HOURS"] != 0]
        title23 = 'W' + Df23['DATETIME'].iloc[0][-2:]
        ax23 = Df23.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title23)
        ax23.grid(color='#988e8c')
        show_values(ax23)

        fill(Df24)
        Df24 = Df24[Df24["TOTAL PROJECT HOURS"] != 0]
        title24 = 'W' + Df24['DATETIME'].iloc[0][-2:]
        ax24 = Df24.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title24)
        ax24.grid(color='#988e8c')
        show_values(ax24)

        fill(Df25)
        Df25 = Df25[Df25["TOTAL PROJECT HOURS"] != 0]
        title25 = 'W' + Df25['DATETIME'].iloc[0][-2:]
        ax25 = Df25.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title25)
        ax25.grid(color='#988e8c')
        show_values(ax25)

        fill(Df26)
        Df26 = Df26[Df26["TOTAL PROJECT HOURS"] != 0]
        title26 = 'W' + Df26['DATETIME'].iloc[0][-2:]
        ax26 = Df26.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title26)
        ax26.grid(color='#988e8c')
        show_values(ax26)

        fill(Df27)
        Df27 = Df27[Df27["TOTAL PROJECT HOURS"] != 0]
        title27 = 'W' + Df27['DATETIME'].iloc[0][-2:]
        ax27 = Df27.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title27)
        ax27.grid(color='#988e8c')
        show_values(ax27)

        fill(Df28)
        Df28 = Df28[Df28["TOTAL PROJECT HOURS"] != 0]
        title28 = 'W' + Df28['DATETIME'].iloc[0][-2:]
        ax28 = Df28.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title28)
        ax28.grid(color='#988e8c')
        show_values(ax28)

        fill(Df29)
        Df29 = Df29[Df29["TOTAL PROJECT HOURS"] != 0]
        title29 = 'W' + Df29['DATETIME'].iloc[0][-2:]
        ax29 = Df29.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title29)
        ax29.grid(color='#988e8c')
        show_values(ax29)

        fill(Df30)
        Df30 = Df30[Df30["TOTAL PROJECT HOURS"] != 0]
        title30 = 'W' + Df30['DATETIME'].iloc[0][-2:]
        ax30 = Df30.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title30)
        ax30.grid(color='#988e8c')
        show_values(ax30)

        fill(Df31)
        Df31 = Df31[Df31["TOTAL PROJECT HOURS"] != 0]
        title31 = 'W' + Df31['DATETIME'].iloc[0][-2:]
        ax31 = Df31.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title31)
        ax31.grid(color='#988e8c')
        show_values(ax31)

        fill(Df32)
        Df32 = Df32[Df32["TOTAL PROJECT HOURS"] != 0]
        title32 = 'W' + Df32['DATETIME'].iloc[0][-2:]
        ax32 = Df32.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title32)
        ax32.grid(color='#988e8c')
        show_values(ax32)

        fill(Df33)
        Df33 = Df33[Df33["TOTAL PROJECT HOURS"] != 0]
        title33 = 'W' + Df33['DATETIME'].iloc[0][-2:]
        ax33 = Df33.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title33)
        ax33.grid(color='#988e8c')
        show_values(ax33)

        fill(Df34)
        Df34 = Df34[Df34["TOTAL PROJECT HOURS"] != 0]
        title34 = 'W' + Df34['DATETIME'].iloc[0][-2:]
        ax34 = Df34.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title34)
        ax34.grid(color='#988e8c')
        show_values(ax34)

        fill(Df35)
        Df35 = Df35[Df35["TOTAL PROJECT HOURS"] != 0]
        title35 = 'W' + Df35['DATETIME'].iloc[0][-2:]
        ax35 = Df35.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title35)
        ax35.grid(color='#988e8c')
        show_values(ax35)

        fill(Df36)
        Df36 = Df36[Df36["TOTAL PROJECT HOURS"] != 0]
        title36 = 'W' + Df36['DATETIME'].iloc[0][-2:]
        ax36 = Df36.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title36)
        ax36.grid(color='#988e8c')
        show_values(ax36)

        fill(Df37)
        Df37 = Df37[Df37["TOTAL PROJECT HOURS"] != 0]
        title37 = 'W' + Df37['DATETIME'].iloc[0][-2:]
        ax37 = Df37.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title37)
        ax37.grid(color='#988e8c')
        show_values(ax37)

        fill(Df38)
        Df38 = Df38[Df38["TOTAL PROJECT HOURS"] != 0]
        title38 = 'W' + Df38['DATETIME'].iloc[0][-2:]
        ax38 = Df38.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title38)
        ax38.grid(color='#988e8c')
        show_values(ax38)

        fill(Df39)
        Df39 = Df39[Df39["TOTAL PROJECT HOURS"] != 0]
        title39 = 'W' + Df39['DATETIME'].iloc[0][-2:]
        ax39 = Df39.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title39)
        ax39.grid(color='#988e8c')
        show_values(ax39)

        fill(Df40)
        Df40 = Df40[Df40["TOTAL PROJECT HOURS"] != 0]
        title40 = 'W' + Df40['DATETIME'].iloc[0][-2:]
        ax40 = Df40.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title40)
        ax40.grid(color='#988e8c')
        show_values(ax40)

        fill(Df41)
        Df41 = Df41[Df41["TOTAL PROJECT HOURS"] != 0]
        title41 = 'W' + Df41['DATETIME'].iloc[0][-2:]
        ax41 = Df41.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title41)
        ax41.grid(color='#988e8c')
        show_values(ax41)

        fill(Df42)
        Df42 = Df42[Df42["TOTAL PROJECT HOURS"] != 0]
        title42 = 'W' + Df42['DATETIME'].iloc[0][-2:]
        ax42 = Df42.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title42)
        ax42.grid(color='#988e8c')
        show_values(ax42)

        fill(Df43)
        Df43 = Df43[Df43["TOTAL PROJECT HOURS"] != 0]
        title43 = 'W' + Df43['DATETIME'].iloc[0][-2:]
        ax43 = Df43.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title43)
        ax43.grid(color='#988e8c')
        show_values(ax43)

        fill(Df44)
        Df44 = Df44[Df44["TOTAL PROJECT HOURS"] != 0]
        title44 = 'W' + Df44['DATETIME'].iloc[0][-2:]
        ax44 = Df44.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title44)
        ax44.grid(color='#988e8c')
        show_values(ax44)

        fill(Df45)
        Df45 = Df45[Df45["TOTAL PROJECT HOURS"] != 0]
        title45 = 'W' + Df45['DATETIME'].iloc[0][-2:]
        ax45 = Df45.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title45)
        ax45.grid(color='#988e8c')
        show_values(ax45)

        fill(Df46)
        Df46 = Df46[Df46["TOTAL PROJECT HOURS"] != 0]
        title46 = 'W' + Df46['DATETIME'].iloc[0][-2:]
        ax46 = Df46.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title46)
        ax46.grid(color='#988e8c')
        show_values(ax46)

        fill(Df47)
        Df47 = Df47[Df47["TOTAL PROJECT HOURS"] != 0]
        title47 = 'W' + Df47['DATETIME'].iloc[0][-2:]
        ax47 = Df47.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title47)
        ax47.grid(color='#988e8c')
        show_values(ax47)

        fill(Df48)
        Df48 = Df48[Df34["TOTAL PROJECT HOURS"] != 0]
        title48 = 'W' + Df48['DATETIME'].iloc[0][-2:]
        ax48 = Df48.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title48)
        ax48.grid(color='#988e8c')
        show_values(ax48)

        fill(Df49)
        Df49 = Df49[Df49["TOTAL PROJECT HOURS"] != 0]
        title49 = 'W' + Df49['DATETIME'].iloc[0][-2:]
        ax49 = Df49.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title49)
        ax49.grid(color='#988e8c')
        show_values(ax49)

        fill(Df50)
        Df50 = Df50[Df50["TOTAL PROJECT HOURS"] != 0]
        title50 = 'W' + Df50['DATETIME'].iloc[0][-2:]
        ax50 = Df50.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title50)
        ax50.grid(color='#988e8c')
        show_values(ax50)

        fill(Df51)
        Df51 = Df51[Df51["TOTAL PROJECT HOURS"] != 0]
        title51 = 'W' + Df51['DATETIME'].iloc[0][-2:]
        ax51 = Df51.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title51)
        ax51.grid(color='#988e8c')
        show_values(ax51)

        fill(Df52)
        Df52 = Df52[Df52["TOTAL PROJECT HOURS"] != 0]
        title52 = 'W' + Df52['DATETIME'].iloc[0][-2:]
        ax52 = Df52.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title52)
        ax52.grid(color='#988e8c')
        show_values(ax52)

        fill(Df53)
        Df53 = Df53[Df53["TOTAL PROJECT HOURS"] != 0]
        title53 = 'W' + Df53['DATETIME'].iloc[0][-2:]
        ax53 = Df53.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title53)
        ax53.grid(color='#988e8c')
        show_values(ax53)

        fill(Df54)
        Df54 = Df54[Df54["TOTAL PROJECT HOURS"] != 0]
        title54 = 'W' + Df54['DATETIME'].iloc[0][-2:]
        ax54 = Df54.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
            kind='bar', rot=10, stacked=True, grid=True, title=title54)
        ax54.grid(color='#988e8c')
        show_values(ax54)
    except:
        pass

    for Df_x in dataframes:
        if Df.empty:
            Df = Df_x
        else:
            Df = pd.concat([Df, Df_x])

    ax = Df.groupby(['CUSTOMER', 'PROJECT NUMBER/ORDER NUMBER'])['TOTAL PROJECT HOURS'].sum().unstack().plot(
        kind='bar', rot=10, stacked=True, grid=True, title='Amount of hours in %s-%s' % (str(start), str(stop)))
    #(str('W' + Df1['DATETIME'].iloc[0][-2:]), str('W' + Df1['DATETIME'].iloc[0][-2:])))
    show_values(ax)
    plt.legend(loc=1)
    plt.grid(True, color='#988e8c')
    plt.show()

if __name__ == "__main__":
    path = dropbox_connect(main_menu())
    create_dataframe(path)
